# Flask imports
from flask import Flask
from flask import abort, redirect, session, flash, g
from flask import render_template, url_for
from flask import request
from flask import send_file
from flask import Markup
from werkzeug import secure_filename

# Python imports
from datetime import datetime
from functools import wraps
from markdown import markdown
from itertools import ifilterfalse
import os
import glob
import random
import sha
import shutil
import smtplib
import subprocess
import tarfile
import zipfile
import sys
try:
	import ldap
except:
	print "ldap not loaded"
from email.mime.text import MIMEText

# Pyboard imports
import serverconfig

# import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)
serverconfig.configure(app)

class User:
	def __init__(self, username):
		self.username = username.strip()
		self.email = None
		self.section = None
		self.grades = []
		self.total = ('', 0, 1)

		try:
			temp = open('users/' + username)
		except:
			flash('Error opening user file.')
			abort(500)
		else:
			for line in temp:
				# ignore blank/invalid lines
				line = line.strip()
				if not line: continue
				if ' ' not in line: continue

				key, val = line.split(' ', 1)
				if key == 'email':
					self.email = val
				elif key == 'section':
					self.section = val
				elif key == 'grade':
					self.grades.append(val.split('\t'))

			temp.close()

		if not self.email or not self.section:
			flash('The user file "' + self.username + '" is not configured properly.')
			abort(500)

		if len(self.grades) > 0:
			self.grades = sorted(self.grades, key = lambda x: x[0])
			self.grades = map(lambda x: (x[0], float(x[1]), float(x[2])), self.grades)
			self.total = reduce(lambda x, y: ('', x[1] + y[1], x[2] + y[2]), self.grades)

	def write(self):
		try:
			temp = open('users/' + self.username, 'w')
		except:
			flash('Error opening user file.')
			abort(500)
		else:
			temp.write('email {}\n'.format(self.email))
			temp.write('section {}\n'.format(self.section))
			for grade in self.grades:
				temp.write('grade {}\t{}\t{}\n'.format(*grade))

			temp.close()

def send_email(me = None, you = None, subject = 'Notification', body = None):
	# TODO send multipart messages instead of only HTML
	# see http://stackoverflow.com/questions/882712/sending-html-email-in-python#882770
	if not me:
		me = app.config['EMAIL_FROM']
	if not you or not body:
		return

	if type(you) is str or type(you) is unicode:
		you = [you]

	message = MIMEText(body, 'html')
	message['Subject'] = subject
	message['From'] = me
	message['To'] = ', '.join(you)

	s = smtplib.SMTP('localhost')
	s.sendmail(me, you, message.as_string())
	s.quit()

def requires_login(func):
	@wraps(func)
	def wrapper(*args, **kwargs):
		if 'username' not in session:
			flash('You must be logged in to view this page.')
			return render_template('login.html')

		return func(*args, **kwargs)
	return wrapper

def requires_admin(func):
	@wraps(func)
	def wrapper(*args, **kwargs):
		if 'username' not in session:
			flash('You must be logged in to view this page.')
			return render_template('login.html')

		if g.user.section != 'admin':
			flash('You must be an administrator to view this page.')
			return redirect(url_for('index'))

		return func(*args, **kwargs)
	return wrapper

def user_in_system(username):
	users = os.listdir('users')
	return username in users

def validate_login(username, password):
	# TODO ldap doesn't like empty passwords... find out why
	if password == '':
		password = ' '
	server = app.config['LDAP_SERVER']
	con = ldap.initialize(server)
	dn = app.config['LDAP_DN'](username)
	rv = con.simple_bind(dn, password)

	try:
		r = con.result(rv)
		if r[0] == 97:
			return True
	except:
		pass

	return False

def handle_file(file_path):
	return_string = ""

	dir = os.path.dirname(file_path)
	ufile = os.path.basename(file_path)

	ext = ufile.split(".")[1:]
	ext = ".".join(ext)

	pdir = os.path.abspath(os.curdir)

	os.chdir(dir)

	if ext == "tar" or ext == "tar.gz":
		return_string += "tar or tar.gz file extension found\n"
		try:
			tf = tarfile.open(ufile)
			tf.extractall()
		except:
			return_string += "ERROR: could not open tar(.gz) file " + ufile + "\n"
			os.chdir(pdir)
			return return_string
		os.remove(ufile)
	elif ext == "zip":
		return_string += "zip file extension found\n"
		try:
			zf = zipfile.ZipFile(ufile)
			zf.extractall()
		except:
			return_string += "ERROR: could not open zip file " + ufile + "\n"
			os.chdir(pdir)
			return return_string
		os.remove(ufile)
	else:
		return_string += "Can't work with " + ufile + ". Don't know how.\n"
		os.chdir(pdir)
		return return_string

	return_string += "file successfully extracted\n"

	find = subprocess.Popen('find .', shell=True, stdout=subprocess.PIPE)
	return_string += "-\n" + find.communicate()[0].strip() + "\n-\n"

	os.mkdir('bin')
	return_string += "---running javac---\n"
	java = subprocess.Popen('javac -d bin -cp .:/usr/share/java/junit.jar **/*.java', shell=True, executable="/bin/zsh", stdout=subprocess.PIPE, stderr=subprocess.PIPE)
	jout = java.communicate()
	if jout[0] or jout[1]:
		return_string += "---javac error/warning output---\n"
	else:
		return_string += "---compilation successful---\n"
	if jout[0]:
		return_string += jout[0] + "\n"
	if jout[1]:
		return_string += jout[1] + "\n"


	os.chdir(pdir)

	return return_string

usr_dir = '/home/pdexter1/public_html/chalk/users/'
xlsx_file = '/home/pdexter1/tmp/grade.xlsx' # this is xlsx file
backup_dir = '/home/pdexter1/backup/' # this is dir which backup the previous record
output_dir = '/home/pdexter1/public_html/chalk/users/'

class quiz_name_Error(Exception):
	pass
class grade_row_Error(Exception):
	pass
class section_row_Error(Exception):
	pass

# Remove hidden files in file list, author JP
def rm_hidden_file(file_list):
    new_list = []
    for file_name in file_list:
	if('~' not in file_name and '.' not in file_name):
	    new_list.append(file_name)
    return new_list

# Generate xlsx file from students file, author JP
def to_xlsx():
    sections = {}  # key is section number, value is worksheet obj
    id_files = rm_hidden_file(os.listdir(usr_dir)) # remove files contain '.' or '~' in name
    sec_quiz = {} # key is section number, value is a dictionary whose key is quiz name, value is column number
    id_names = {} # key is section number, value is names list in that section

    workbook = Workbook()

    for id_file in id_files:
        f = open(usr_dir + id_file, 'r')
        for line in f:
            item = line.split()
            # check user_id and email
            if(item[0] == 'email'):
                continue
            # class section
            elif(item[0] == 'section'):
		if(len(item) != 2):
			raise section_row_Error
                sec_name = item[1]
                if(sec_name not in sections): # if section number not in sections, then create a new sheet for section
                    worksheet = workbook.create_sheet(0)
                    worksheet.title = sec_name
                    sections[sec_name] = worksheet
                    sec_quiz[sec_name] = {}
                else:
                    worksheet = sections[sec_name]
                if(sec_name not in id_names):
                    id_names[sec_name] = [id_file]
                else:
                    id_names[sec_name].append(id_file)
            # fill in grade
            elif(item[0] == 'grade'):
		if(len(item) != 4):
			raise grade_row_Error
                quiz_name = item[1]
                grade = item[2]
		max_score = item[3]
                if(quiz_name not in sec_quiz[sec_name]):
                    sec_quiz[sec_name][quiz_name] = len(sec_quiz[sec_name]) + 1
                    worksheet.cell(row = 0, column = sec_quiz[sec_name][quiz_name]).value = quiz_name + ' ' + max_score
                worksheet.cell(row = len(id_names[sec_name]), column = sec_quiz[sec_name][quiz_name]).value = grade
            # error
            else:
                print 'Should have no field called ' + item[0] #TODO it should goes to logs
                id_file = ''
                f.close()
                break

        if(id_file):
            worksheet.cell(row = len(id_names[sec_name]), column = 0).value = id_file # fill in id
        f.close()

    workbook.save(xlsx_file)

# Parse xlsx to grade files, author JP
def from_xlsx():

    wb = load_workbook(xlsx_file)
    section_names = wb.get_sheet_names()

    for section_name in section_names:
        ws = wb.get_sheet_by_name(section_name)
        end_row = ws.get_highest_row()
        end_column = ws.get_highest_column()
        quiz_names = []
        for quiz_column in range(1, end_column):
            quiz_names.append(str(ws.cell(row = 0, column = quiz_column).value))

        for name_row in range(1, end_row):
            file_name = str(ws.cell(row = name_row, column = 0).value)
            f = open(output_dir + file_name, 'w')
            f.write('email ' + file_name + '@binghamton.edu\n')
            f.write('section ' + section_name + '\n')
            for quiz_grade in range(1, len(quiz_names) + 1):
                grade = str(ws.cell(row = name_row, column = quiz_grade).value)
                name_max_score = quiz_names[quiz_grade - 1].split()
		if(len(name_max_score) != 2):
			raise quiz_name_Error
                f.write('grade ' + name_max_score[0] + '\t' + grade + '\t' + name_max_score[1] + '\n')
            f.close()

# Download the grades, author JP
@app.route('/admin/getexcel')
@requires_admin
def get_excel_file():
	try:
		to_xlsx()
		return send_file(xlsx_file, as_attachment=True)
	except section_row_Error:
		e = 'The section format should be "section [section code]" '
		return render_template('fail.html', ErrorType = e)
	except grade_row_Error:
		e = 'The grade format should be "grade [quiz name]	[grade]		[max grade]" '
		return render_template('fail.html', ErrorType = e)
	except:
		e = sys.exc_info()[0]
		return render_template('fail.html', ErrorType = e)

# Upload the grades, author JP
@app.route('/admin/putexcel', methods=['GET', 'POST'])
@requires_admin
def put_excel_file():
	try:
		if request.method == 'POST':
			user = User(session['username'])
			ufile = request.files['file']
			filename = secure_filename(ufile.filename)

			d = datetime.now()
			time_stamp = d.strftime("%Y-%m-%d-%H-%M-%S")
			os.system("tar czPf " + backup_dir + time_stamp + ".tar.gz " + output_dir)

			ufile.save(xlsx_file)
			from_xlsx()
			return render_template('upload.html', title='Upload',
				name = user.username,
				filename = filename,
				section = user.section,
				assignment = 'all assignment',
				output = 'The grade is overloaded by uploaded xlsx file')
		else:
			return render_template('upload_grade.html', title='Submit')
	except quiz_name_Error:
		e = 'The quiz name format should be "[quiz name] [max score]" '
		return render_template('fail.html', ErrorType = e)
	except:
		e = sys.exc_info()[0]
		return render_template('fail.html', ErrorType = e)

def get_submissions(section, assignment):
	pdir = os.path.abspath(os.curdir)
	dir = "files/" + section + "/" + assignment + "/"
	try:
		os.chdir(dir)
	except:
		os.chdir(pdir)
		return None
	os.chdir('..')
	file_name = section + "_" + assignment + ".tar.gz"
	tar = tarfile.open(file_name, "w:gz")
	tar.add(assignment)
	tar.close()

	os.chdir(pdir)

	shutil.move(dir + '../' + file_name, "static/" + file_name)
	return "static/" + file_name

def grades_for_user(username):
	user = User(username)
	if len(user.grades) == 0:
		flash('No grades to view')
	return render_template('grades.html',
		grades=user.grades,
		total=user.total)

def retrieve_users():
	return [User(user) for user in ifilterfalse(lambda x: x.startswith('.'), sorted(os.listdir('users')))]

@app.before_request
def before():
	if 'username' in session:
		g.user = User(session['username'])

@app.route('/', methods=['GET', 'POST'])
@requires_login
def index():
	if request.method == 'POST':
		user = User(session['username'])

		ufile = request.files['file']
		filename = secure_filename(ufile.filename)
		assignment = request.form['ass']
		d = datetime.now()
		time_stamp = d.strftime("%Y-%m-%d-%H-%M-%S")
		upload_dir = "files/" + user.section + "/" + assignment + "/" + user.username + "/" + time_stamp + "/"

		try:
			os.makedirs(upload_dir)
		except:
			pass # uhh wat FIXME

		full_path = upload_dir + filename
		ufile.save(full_path)

		num_submission = len(glob.glob('files/'+user.section+'/'+assignment+'/'+user.username+'/*'))
		verification_code = sha.new(app.config['VERIFICATION_SALT'] + user.username + assignment + str(num_submission)).hexdigest()
		# FIXME template
		send_email(you = user.email,
			subject = 'Submission Received',
			body = render_template('email_submission.html',
				verification = verification_code,
				assignment = assignment))

		output = handle_file(full_path)
		return render_template('upload.html', title='Upload',
			name = user.username,
			filename = filename,
			section = user.section,
			assignment = assignment,
			output = output)
	else:
		assignments = [line.strip() for line in open('assignments') if line.strip()]
		assignments.reverse()
		return render_template('index.html', title='Submit',
			assignments = assignments)

@app.route('/mysubmissions')
@requires_login
def view_submissions():
	assignment = request.args.get('assignment')
	timestamp = request.args.get('timestamp')
	username = session['username']
	user = User(username)
	if assignment is None or timestamp is None:
		submissions = glob.glob('files/'+user.section+'/**/'+user.username)
		if len(submissions) == 0:
			flash('No submissions found')
		else:
			submissions = map(lambda x: (x.split('/')[2], os.listdir(x)), submissions)
		return render_template('submissions.html', submissions=submissions)
	else:
		pdir = os.path.abspath(os.curdir)
		dir = "files/" + user.section + "/" + assignment + "/" + user.username + "/" + timestamp
		try:
			os.chdir(dir)
		except:
			os.chdir(pdir)
			return abort(404)
		os.chdir('../..')
		file_name = user.username + "_" + timestamp + ".tar.gz"
		tar = tarfile.open(file_name, "w:gz")
		tar.add(user.username + '/' + timestamp)
		tar.close()
		os.chdir(pdir)
		shutil.move(dir + '/../../' + file_name, "static/" + file_name)
		return send_file("static/" + file_name, as_attachment=True)

@app.route('/download', methods=['GET', 'POST'])
@requires_admin
def download_admin():
	if request.method == 'GET':
		assignments = [line.strip() for line in open('assignments') if line.strip()]
		assignments.reverse()
		return render_template('download.html',
			assignments=assignments)
	else:
		file_name = get_submissions(request.form['section'], request.form['ass'])
		if file_name is None:
			flash('Section or Assignment not found')
			return abort(404)
		return send_file(file_name, as_attachment=True)

@app.route('/announcements')
@requires_login
def announcements():
	announcements = [markdown(open('announcements/'+a).read()) for a in sorted(os.listdir('announcements'))]
	announcements.reverse()
	return render_template('announcements.html', announcements=announcements)

@app.route('/admin/announcements', methods=['GET', 'POST'])
@requires_admin
def announcements_admin():
	if request.method == 'GET':
		return render_template('announcements_create.html')
	else:
		announcement = request.form['announcement']
		timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

		temp = open('announcements/' + timestamp, 'w')
		temp.write(announcement)
		temp.close()

		yous = []
		users = retrieve_users()
		for user in users:
			yous.append(user.email)

		send_email(you = yous,
			subject = 'New announcement for ' + app.config['COURSE_NAME'],
			body = render_template('email_announcement.html',
				announcement = markdown(announcement)))

		return redirect(url_for('announcements'))

@app.route('/admin')
@requires_admin
def admin():
	return render_template('admin.html')

@app.route('/profile', methods=['GET', 'POST'])
@app.route('/user/<username>')
@requires_login
def profile(username = None):
	if not username:
		username = session['username']
	try:
		user = User(username)
	except:
		return abort(404)

	if request.method == 'POST':
		user.email = request.form['new_email']
		user.write()

		flash('Email updated to "' + user.email + '"')

	return render_template('profile.html',
		user = user,
		isuser = 'username' in session and session['username'] == username)

@app.route('/profile/edit')
@requires_login
def profile_edit():
	username = session['username']
	try:
		user = User(username)
	except:
		return abort(404)

	return render_template('profile_edit.html',
		user = user)

@app.route('/users')
@requires_login
def users():
	users = retrieve_users()
	return render_template('users.html',
		title='Email List',
		users=users)

@app.route('/grades')
@requires_login
def grades():
	return grades_for_user(session['username'])

@app.route('/admin/grades')
@app.route('/admin/grades/<username>')
@requires_admin
def grades_admin(username = None):
	if username is None:
		users = retrieve_users()
		users = [user for user in users if user.section != 'admin']
		return render_template('class_grades.html',
			title = 'Class Grades',
			users = users)
	else:
		return grades_for_user(username)

@app.route('/login', methods=['GET', 'POST'])
def login():
	if request.method == 'GET':
		return render_template('login.html')
	else:
		if validate_login(request.form['username'], request.form['password']):
			if not user_in_system(request.form['username']):
				flash('User not in system')
				return render_template('login.html')
			session['username'] = request.form['username']
			flash('Logged in')
			return redirect(url_for('index'))

		flash('Invalid login')
		return redirect(url_for('login'))

@app.route('/logout')
def logout():
	session.pop('username', None)
	flash('Logged out')
	return redirect(url_for('login'))

@app.errorhandler(400)
@app.errorhandler(401)
@app.errorhandler(403)
@app.errorhandler(404)
@app.errorhandler(500)
def view_error(error):
	try:
		return render_template('error.html',
			error = '{} Error'.format(error.code),
			desc = error.description), error.code

	except:
		return render_template('error.html',
			error = 'Oh my',
			desc = 'Something went wrong.')

if __name__ == '__main__':
	app.run(host='0.0.0.0', debug=True)
